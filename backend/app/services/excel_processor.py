"""Excel file processing service for BBG rebate files."""
from typing import Optional, Dict, Any, List, Tuple
from pathlib import Path
import openpyxl

from app.utils.exceptions import ExcelProcessingError


class ExcelProcessor:
    """Processes BBG rebate Excel files (.xlsm) in streaming (read-only) mode."""

    # Keywords to detect header row
    HEADER_KEYWORDS = [
        "date", "job code", "job name", "address", "city",
        "state", "zip", "postal", "product", "qty", "quantity"
    ]

    # Number of leading rows to cache for random-access parsing (metadata,
    # header detection, active-product identification).
    _TOP_ROW_CACHE_SIZE = 30

    def __init__(self, file_path: str):
        """Initialize processor with file path.

        Args:
            file_path: Path to the .xlsm file
        """
        self.file_path = Path(file_path)
        self.workbook: Optional[openpyxl.Workbook] = None
        self.reformatter_sheet_name: Optional[str] = None
        self.programs_products_sheet_name: Optional[str] = None
        self.metadata: Dict[str, Any] = {}
        self.file_products: Dict[str, Dict[str, Any]] = {}
        self.headers: List[Any] = []  # Header row values, for data_transformer

        # Cache of first N rows of Usage-Reporting sheet (as tuples of values).
        # Populated by find_usage_reporting_sheet. Enables random-access parsing
        # of metadata / header / product rows without leaving streaming mode.
        self._top_rows: List[Tuple] = []

    def open_file(self) -> None:
        """Open the Excel file and validate it exists."""
        if not self.file_path.exists():
            raise ExcelProcessingError(f"File not found: {self.file_path}")

        if not self.file_path.suffix.lower() in ['.xlsm', '.xlsx']:
            raise ExcelProcessingError(
                f"Invalid file type: {self.file_path.suffix}. Expected .xlsm or .xlsx"
            )

        try:
            # read_only=True streams rows instead of materializing a Python Cell
            # object for every cell in the sheet's used range. Essential for
            # files whose used range is inflated (e.g. formatting extending to
            # Excel's 1,048,576-row limit) -- those would otherwise consume
            # hundreds of MB to GB at open time.
            self.workbook = openpyxl.load_workbook(
                self.file_path,
                data_only=True,
                read_only=True,
            )
        except Exception as e:
            raise ExcelProcessingError(f"Failed to open Excel file: {str(e)}")

    @property
    def reformatter_sheet(self):
        """The Usage-Reporting sheet (read-only, for streaming iteration)."""
        if not self.workbook or not self.reformatter_sheet_name:
            return None
        return self.workbook[self.reformatter_sheet_name]

    @property
    def programs_products_sheet(self):
        """The Programs-Products sheet (read-only), if found."""
        if not self.workbook or not self.programs_products_sheet_name:
            return None
        return self.workbook[self.programs_products_sheet_name]

    def find_usage_reporting_sheet(self) -> None:
        """Locate the 'Usage-Reporting' sheet and cache its top rows."""
        if not self.workbook:
            raise ExcelProcessingError("Workbook not loaded. Call open_file() first.")

        for sheet_name in self.workbook.sheetnames:
            sheet_lower = sheet_name.lower()
            if 'usage' in sheet_lower and 'report' in sheet_lower:
                self.reformatter_sheet_name = sheet_name
                ws = self.workbook[sheet_name]
                self._top_rows = list(
                    ws.iter_rows(max_row=self._TOP_ROW_CACHE_SIZE, values_only=True)
                )
                return

        raise ExcelProcessingError(
            "Usage-Reporting sheet not found. Expected a sheet with "
            "'usage' and 'report' in the name."
        )

    def _get_cell(self, row: int, col: int) -> Any:
        """Get a cell's value from the cached top rows (1-indexed)."""
        if row < 1 or row > len(self._top_rows):
            return None
        row_tuple = self._top_rows[row - 1]
        if col < 1 or col > len(row_tuple):
            return None
        return row_tuple[col - 1]

    def extract_metadata(self) -> Dict[str, str]:
        """Extract metadata from cells B6 (bbg_member_id) and B7 (member_name).

        Supports both OLD and NEW formats:
        - NEW: B6 has member ID, B7 has member name
        - OLD: B6 is blank, B7 has member name (ID must be looked up)

        Returns:
            Dictionary with bbg_member_id and member_name (ID may be None for old format)
        """
        if not self._top_rows:
            raise ExcelProcessingError("Usage-Reporting sheet not loaded.")

        bbg_member_id = self._get_cell(6, 2)  # B6
        member_name = self._get_cell(7, 2)    # B7

        if not member_name:
            raise ExcelProcessingError("Cell B7 (Member Name) is empty")

        if not bbg_member_id:
            # Old format - ID will be looked up by name during enrichment.
            self.metadata = {
                'bbg_member_id': None,
                'member_name': str(member_name).strip(),
                'format': 'OLD'
            }
        else:
            self.metadata = {
                'bbg_member_id': str(bbg_member_id).strip(),
                'member_name': str(member_name).strip(),
                'format': 'NEW'
            }

        return self.metadata

    def detect_header_row(self, max_rows: int = 20) -> int:
        """Auto-detect the header row by searching for keywords.

        Args:
            max_rows: Maximum number of rows to search

        Returns:
            Row number (1-indexed) where headers are found
        """
        if not self._top_rows:
            raise ExcelProcessingError("Reformatter sheet not loaded.")

        limit = min(max_rows, len(self._top_rows))
        for row_num in range(1, limit + 1):
            row = self._top_rows[row_num - 1]
            row_values = [str(v).lower() for v in row if v is not None and v != '']

            matches = sum(
                1 for keyword in self.HEADER_KEYWORDS
                if any(keyword in val for val in row_values)
            )

            if matches >= 3:
                # Cache header values so the transformer doesn't need random
                # cell access on the (streaming-only) worksheet.
                self.headers = list(row)
                return row_num

        raise ExcelProcessingError(
            f"Header row not found in first {max_rows} rows. "
            f"Expected keywords: {', '.join(self.HEADER_KEYWORDS[:5])}"
        )

    def identify_active_products(self, header_row: int) -> Dict[int, Dict[str, str]]:
        """Identify active product columns and extract product IDs and distributor info.

        Args:
            header_row: The row number where headers are located

        Returns:
            Dictionary mapping column index to dict with product_id and distributor
        """
        if not self._top_rows:
            raise ExcelProcessingError("Usage-Reporting sheet not loaded.")

        active_products: Dict[int, Dict[str, str]] = {}

        # Row 2 = active flags, row 5 = distributor names, row 7 = product IDs.
        row_2 = self._top_rows[1] if len(self._top_rows) > 1 else ()
        row_5 = self._top_rows[4] if len(self._top_rows) > 4 else ()
        row_7 = self._top_rows[6] if len(self._top_rows) > 6 else ()

        max_col = max(len(row_2), len(row_5), len(row_7))
        for col_idx in range(1, max_col + 1):
            active_flag = row_2[col_idx - 1] if col_idx <= len(row_2) else None
            distributor = row_5[col_idx - 1] if col_idx <= len(row_5) else None
            product_id = row_7[col_idx - 1] if col_idx <= len(row_7) else None

            is_active = (
                active_flag == 1
                or active_flag == 1.0
                or str(active_flag).strip() == '1'
            )

            if is_active and product_id:
                try:
                    if isinstance(product_id, (int, float)):
                        pid_str = str(int(product_id))
                    elif str(product_id).isdigit():
                        pid_str = str(product_id).strip()
                    else:
                        continue  # Skip non-numeric product IDs

                    active_products[col_idx] = {
                        'product_id': pid_str,
                        'distributor': str(distributor).strip() if distributor else None
                    }
                except Exception:
                    pass  # Skip invalid entries

        if not active_products:
            raise ExcelProcessingError(
                "No active products found. Expected Row 2 = 1 for active columns "
                "and numeric product IDs in Row 7."
            )

        return active_products

    def extract_programs_products(self) -> Dict[str, Dict[str, Any]]:
        """Extract Programs-Products data from the Excel file.

        Returns:
            Dictionary mapping product_id to product info (program, name, proof_point)
        """
        if not self.workbook:
            raise ExcelProcessingError("Workbook not loaded.")

        programs_sheet = None
        for sheet_name in self.workbook.sheetnames:
            if 'program' in sheet_name.lower() and 'product' in sheet_name.lower():
                self.programs_products_sheet_name = sheet_name
                programs_sheet = self.workbook[sheet_name]
                break

        if not programs_sheet:
            return {}

        products: Dict[str, Dict[str, Any]] = {}

        # Stream rows; row 1 = headers, row 2+ = data.
        # Columns: A (program), B (product_id), C (product_name), D (proof_point).
        for row_idx, row in enumerate(programs_sheet.iter_rows(values_only=True), start=1):
            if row_idx == 1:
                continue

            program = row[0] if len(row) > 0 else None
            product_id = row[1] if len(row) > 1 else None
            product_name = row[2] if len(row) > 2 else None
            proof_point = row[3] if len(row) > 3 else None

            # Skip empty rows
            if not product_id or not product_name:
                continue

            # Skip repeated header rows
            if str(program).strip() in ('Product ID', 'Program'):
                continue

            if isinstance(product_id, (int, float)):
                product_id_str = str(int(product_id))
            else:
                product_id_str = str(product_id).strip()

            products[product_id_str] = {
                'program_name': str(program).strip() if program else None,
                'product_name': str(product_name).strip() if product_name else None,
                'proof_point': str(proof_point).strip() if proof_point and str(proof_point) != 'None' else None,
            }

        self.file_products = products
        return products

    def get_column_letter(self, col_idx: int) -> str:
        """Convert column index to Excel column letter (e.g., 1 -> A, 27 -> AA)."""
        return openpyxl.utils.get_column_letter(col_idx)

    def close(self) -> None:
        """Close the workbook and release resources."""
        if self.workbook:
            self.workbook.close()
            self.workbook = None
            self.reformatter_sheet_name = None
            self.programs_products_sheet_name = None
            self._top_rows = []

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()
