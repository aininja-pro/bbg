"""Service for generating supplier and territory manager distribution reports."""
import io
import re
import gc
import zipfile
from typing import List, Dict, Optional
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select

from app.models.lookup import TradeNetMember


# Column definitions for each mode
MODE1_COLUMNS = [
    'supplier_name', 'address_type', 'address', 'closing_date',
    'product_id', 'product', 'quantity', 'total_rebates',
    'member_name', 'pp_dist_subcontractor', 'pp_prod_purchase',
    'year', 'quarter', 'state'
]

MODE2_COLUMNS = [
    'supplier_name', 'address_type', 'address', 'closing_date', 'category',
    'product_id', 'product', 'quantity', 'bbg_member_id', 'member_name', 'tm',
    'pp_dist_subcontractor', 'pp_prod_purchase', 'year', 'quarter', 'state'
]


class ReportGenerator:
    """Generate supplier and territory manager distribution reports."""

    @staticmethod
    def format_worksheet(ws):
        """
        Format worksheet with bold headers, auto-sized columns, frozen header, and convert text numbers to real numbers.

        Args:
            ws: openpyxl Worksheet object
        """
        # Freeze the header row (row 1)
        ws.freeze_panes = 'A2'

        # Make header row bold and centered
        for cell in ws[1]:
            cell.font = Font(bold=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Auto-size columns and convert numeric strings to numbers
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for idx, cell in enumerate(column):
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length

                        # Convert numeric strings to actual numbers (skip header)
                        if idx > 0 and isinstance(cell.value, str):
                            # Try to convert to number to avoid Excel green warning flags
                            try:
                                # Try integer first
                                if '.' not in cell.value:
                                    cell.value = int(cell.value)
                                else:
                                    cell.value = float(cell.value)
                            except (ValueError, AttributeError):
                                # Keep as string if conversion fails
                                pass
                except:
                    pass

            # Set column width with some padding (max 50 chars)
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    @staticmethod
    def sanitize_sheet_name(name: str) -> str:
        r"""
        Sanitize sheet name for Excel.
        - Remove invalid characters: \/*?[]:
        - Truncate to 31 characters (Excel limit)
        """
        if not name:
            return "Sheet"

        # Remove invalid characters
        sanitized = re.sub(r'[\\/*?[\]:]', '', str(name))

        # Truncate to 31 characters
        sanitized = sanitized[:31]

        # Remove trailing/leading spaces
        sanitized = sanitized.strip()

        return sanitized if sanitized else "Sheet"

    @staticmethod
    def sanitize_filename(name: str) -> str:
        """
        Sanitize filename for file system.
        - Keep only alphanumeric, spaces, hyphens, underscores
        - Replace spaces with underscores
        """
        if not name:
            return "file"

        # Replace spaces with underscores
        sanitized = str(name).replace(' ', '_')

        # Keep only alphanumeric, hyphens, underscores
        sanitized = re.sub(r'[^a-zA-Z0-9_-]', '', sanitized)

        return sanitized if sanitized else "file"

    @staticmethod
    async def merge_csv_files(file_paths: List[str]) -> pd.DataFrame:
        """
        Merge multiple CSV files into a single DataFrame.

        Args:
            file_paths: List of paths to CSV files

        Returns:
            Merged DataFrame
        """
        dfs = []

        for file_path in file_paths:
            try:
                df = pd.read_csv(file_path, low_memory=False)
                dfs.append(df)
            except Exception as e:
                raise ValueError(f"Failed to read {file_path}: {str(e)}")

        if not dfs:
            raise ValueError("No valid CSV files to merge")

        # Concatenate all dataframes
        merged_df = pd.concat(dfs, ignore_index=True)

        # Clean up
        del dfs
        gc.collect()

        return merged_df

    @staticmethod
    async def enrich_with_tm(df: pd.DataFrame, db: AsyncSession) -> pd.DataFrame:
        """
        Enrich DataFrame with Territory Manager data by joining with lookup_members.

        Args:
            df: DataFrame to enrich
            db: Database session

        Returns:
            Enriched DataFrame with 'tm' column populated
        """
        # Fetch all members with TM data from database
        result = await db.execute(
            select(TradeNetMember.bbg_member_id, TradeNetMember.territory_manager)
        )
        members = result.all()

        # Create a lookup dictionary: bbg_member_id -> territory_manager
        tm_lookup = {
            str(member.bbg_member_id): member.territory_manager
            for member in members
            if member.bbg_member_id
        }

        # Add tm column if it doesn't exist
        if 'tm' not in df.columns:
            df['tm'] = None

        # Populate tm column using bbg_member_id lookup
        df['bbg_member_id'] = df['bbg_member_id'].astype(str)
        df['tm'] = df['bbg_member_id'].map(tm_lookup)

        # Handle unassigned TMs (NULL or empty)
        df['tm'] = df['tm'].fillna('Unassigned')
        df['tm'] = df['tm'].replace('', 'Unassigned')

        return df

    @staticmethod
    def create_excel_from_df(df: pd.DataFrame, columns: List[str], add_empty_cols: int = 0) -> bytes:
        """
        Create an Excel workbook from a DataFrame with specified columns.

        Args:
            df: DataFrame to convert
            columns: List of column names to include
            add_empty_cols: Number of empty columns to add at the end

        Returns:
            Excel file as bytes
        """
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Filter DataFrame to only include specified columns
        df_filtered = df[columns].copy() if all(col in df.columns for col in columns) else df

        # Create single sheet
        ws = wb.create_sheet("Data")

        # Write data to sheet
        for r_idx, row in enumerate(dataframe_to_rows(df_filtered, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # Add empty columns if requested
        if add_empty_cols > 0:
            max_col = ws.max_column
            for col_offset in range(1, add_empty_cols + 1):
                ws.cell(row=1, column=max_col + col_offset, value=None)

        # Save to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return buffer.getvalue()

    @staticmethod
    def generate_mode1_excel(df: pd.DataFrame, supplier_name: str) -> bytes:
        """
        Generate Mode 1 Excel: Supplier with Product tabs.

        Structure:
        - Sheet "AllData": All rows for this supplier (16 columns + 2 empty)
        - Sheet per product_id: Rows for that product (14 columns)

        Args:
            df: DataFrame with data for this supplier
            supplier_name: Name of the supplier

        Returns:
            Excel file as bytes
        """
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Filter to relevant columns
        df_filtered = df[MODE1_COLUMNS].copy() if all(col in df.columns for col in MODE1_COLUMNS) else df

        # Sort by member_name, state, city, zip_postal
        sort_columns = []
        if 'member_name' in df_filtered.columns:
            sort_columns.append('member_name')
        if 'state' in df_filtered.columns:
            sort_columns.append('state')
        if 'city' in df_filtered.columns:
            sort_columns.append('city')
        if 'zip_postal' in df_filtered.columns:
            sort_columns.append('zip_postal')

        if sort_columns:
            df_filtered = df_filtered.sort_values(sort_columns)

        # Sheet 1: "AllData" with all rows (add 2 empty columns)
        ws_all = wb.create_sheet("AllData")
        for r_idx, row in enumerate(dataframe_to_rows(df_filtered, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws_all.cell(row=r_idx, column=c_idx, value=value)

        # Add 2 empty columns
        max_col = ws_all.max_column
        ws_all.cell(row=1, column=max_col + 1, value=None)
        ws_all.cell(row=1, column=max_col + 2, value=None)

        # Format the AllData sheet
        ReportGenerator.format_worksheet(ws_all)

        # Create additional sheets for each unique product_id
        if 'product_id' in df_filtered.columns:
            product_ids = df_filtered['product_id'].unique()

            # Sort product IDs in ascending order for tab creation
            product_ids_sorted = sorted([str(pid) for pid in product_ids if not pd.isna(pid) and pid != ''])

            for product_id in product_ids_sorted:
                if pd.isna(product_id) or product_id == '':
                    continue

                # Filter data for this product
                product_df = df_filtered[df_filtered['product_id'] == product_id].copy()

                # Sanitize sheet name
                sheet_name = ReportGenerator.sanitize_sheet_name(str(product_id))

                # Handle duplicate sheet names
                original_name = sheet_name
                counter = 1
                while sheet_name in wb.sheetnames:
                    sheet_name = f"{original_name[:28]}_{counter}"
                    counter += 1

                # Create sheet
                ws = wb.create_sheet(sheet_name)

                # Write data (without empty columns for product tabs)
                for r_idx, row in enumerate(dataframe_to_rows(product_df, index=False, header=True), 1):
                    for c_idx, value in enumerate(row, 1):
                        ws.cell(row=r_idx, column=c_idx, value=value)

                # Format the product sheet
                ReportGenerator.format_worksheet(ws)

        # Save to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return buffer.getvalue()

    @staticmethod
    def generate_mode2_excel(df: pd.DataFrame, tm_name: str) -> bytes:
        """
        Generate Mode 2 Excel: Territory Manager with Supplier tabs.

        Structure:
        - Sheet "All Data": All rows for this TM (16 columns)
        - Sheet per supplier_name: Rows for that supplier (16 columns)

        Args:
            df: DataFrame with data for this TM
            tm_name: Name of the territory manager

        Returns:
            Excel file as bytes
        """
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Filter to relevant columns
        df_filtered = df[MODE2_COLUMNS].copy() if all(col in df.columns for col in MODE2_COLUMNS) else df

        # Sort by member_name, state, city, zip_postal (same as Mode 1)
        sort_columns = []
        if 'member_name' in df_filtered.columns:
            sort_columns.append('member_name')
        if 'state' in df_filtered.columns:
            sort_columns.append('state')
        if 'city' in df_filtered.columns:
            sort_columns.append('city')
        if 'zip_postal' in df_filtered.columns:
            sort_columns.append('zip_postal')

        if sort_columns:
            df_filtered = df_filtered.sort_values(sort_columns)

        # Sheet 1: "All Data" with all rows
        ws_all = wb.create_sheet("All Data")
        for r_idx, row in enumerate(dataframe_to_rows(df_filtered, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws_all.cell(row=r_idx, column=c_idx, value=value)

        # Format the All Data sheet
        ReportGenerator.format_worksheet(ws_all)

        # Create additional sheets for each unique supplier_name
        if 'supplier_name' in df_filtered.columns:
            supplier_names = df_filtered['supplier_name'].unique()

            # Sort supplier names alphabetically (A-Z)
            supplier_names_sorted = sorted([str(name) for name in supplier_names if not pd.isna(name) and name != ''])

            for supplier_name in supplier_names_sorted:
                if pd.isna(supplier_name) or supplier_name == '':
                    continue

                # Filter data for this supplier
                supplier_df = df_filtered[df_filtered['supplier_name'] == supplier_name].copy()

                # Sanitize sheet name
                sheet_name = ReportGenerator.sanitize_sheet_name(str(supplier_name))

                # Handle duplicate sheet names
                original_name = sheet_name
                counter = 1
                while sheet_name in wb.sheetnames:
                    sheet_name = f"{original_name[:28]}_{counter}"
                    counter += 1

                # Create sheet
                ws = wb.create_sheet(sheet_name)

                # Write data
                for r_idx, row in enumerate(dataframe_to_rows(supplier_df, index=False, header=True), 1):
                    for c_idx, value in enumerate(row, 1):
                        ws.cell(row=r_idx, column=c_idx, value=value)

                # Format the supplier sheet
                ReportGenerator.format_worksheet(ws)

        # Save to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return buffer.getvalue()

    @staticmethod
    async def create_distribution_zip(
        mode: str,
        df: pd.DataFrame,
        db: Optional[AsyncSession] = None,
        progress_callback = None
    ) -> bytes:
        """
        Create a ZIP file containing all Excel files for the selected mode.

        Args:
            mode: "mode1" (Supplier with Product tabs) or "mode2" (TM with Supplier tabs)
            df: DataFrame with all data
            db: Database session (required for mode2 to enrich with TM data)

        Returns:
            ZIP file as bytes
        """
        zip_buffer = io.BytesIO()

        # Use ZIP_STORED (no compression) for better compatibility with macOS Archive Utility
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_STORED) as zip_file:
            if mode == "mode1":
                # Mode 1: Generate one Excel per supplier
                if 'supplier_name' not in df.columns:
                    raise ValueError("DataFrame must have 'supplier_name' column for Mode 1")

                supplier_names = df['supplier_name'].unique()
                total_suppliers = len(supplier_names)
                processed_count = 0

                for supplier_name in supplier_names:
                    if pd.isna(supplier_name) or supplier_name == '':
                        continue

                    # Filter data for this supplier
                    supplier_df = df[df['supplier_name'] == supplier_name].copy()

                    # Generate Excel
                    excel_bytes = ReportGenerator.generate_mode1_excel(supplier_df, supplier_name)

                    # Add to ZIP with sanitized filename
                    filename = f"{ReportGenerator.sanitize_filename(str(supplier_name))}.xlsx"
                    zip_file.writestr(filename, excel_bytes)

                    # Update progress
                    processed_count += 1
                    if progress_callback:
                        progress_pct = int((processed_count / total_suppliers) * 100)
                        await progress_callback(progress_pct, f"{processed_count} of {total_suppliers} suppliers")

                    # Clean up
                    del supplier_df, excel_bytes
                    gc.collect()

            elif mode == "mode2":
                # Mode 2: Enrich with TM data, then generate one Excel per TM
                if db is None:
                    raise ValueError("Database session required for Mode 2")

                # Enrich with TM data
                df_enriched = await ReportGenerator.enrich_with_tm(df, db)

                if 'tm' not in df_enriched.columns:
                    raise ValueError("DataFrame must have 'tm' column for Mode 2")

                tm_names = df_enriched['tm'].unique()
                total_tms = len(tm_names)
                processed_count = 0

                for tm_name in tm_names:
                    if pd.isna(tm_name) or tm_name == '':
                        tm_name = 'Unassigned'

                    # Filter data for this TM
                    tm_df = df_enriched[df_enriched['tm'] == tm_name].copy()

                    # Generate Excel
                    excel_bytes = ReportGenerator.generate_mode2_excel(tm_df, tm_name)

                    # Add to ZIP with sanitized filename
                    filename = f"TM_{ReportGenerator.sanitize_filename(str(tm_name))}.xlsx"
                    zip_file.writestr(filename, excel_bytes)

                    # Update progress
                    processed_count += 1
                    if progress_callback:
                        progress_pct = int((processed_count / total_tms) * 100)
                        await progress_callback(progress_pct, f"{processed_count} of {total_tms} TMs")

                    # Clean up
                    del tm_df, excel_bytes
                    gc.collect()

                # Clean up enriched DataFrame
                del df_enriched
                gc.collect()

            else:
                raise ValueError(f"Invalid mode: {mode}. Must be 'mode1' or 'mode2'")

        zip_buffer.seek(0)
        return zip_buffer.getvalue()
